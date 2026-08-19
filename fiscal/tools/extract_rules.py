"""Extrai regras fiscais sem misturar empresas.

- OK.xlsx, primeira aba → BAIFER
- ODS aba LOJA → Loja das Máquinas
"""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

from odf.opendocument import load
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import load_workbook

DESTINO_KEYS = [
    "naoContribuinte",
    "contribuinte",
    "revenda",
    "construtora",
    "hospClinica",
    "orgaoPublico",
    "produtorRural",
    "atacado",
]

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OK = ROOT / "data" / "ok-baifer.xlsx"
DEFAULT_ODS = ROOT / "data" / "relacao-produtos-ncms-baifer.ods"
DEFAULT_BAIFER_JSON = ROOT / "data" / "base-baifer.json"
DEFAULT_LOJA_JSON = ROOT / "data" / "base-loja.json"
FORBIDDEN_SHEETS = {"Planilha_Classes_Fiscais"}


def normalize_ncm(raw: object | None) -> str:
    if raw is None:
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return ""
    if len(digits) < 8:
        return digits.zfill(8)
    return digits[:8]


def _strip_accents(text: str) -> str:
    nfd = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in nfd if unicodedata.category(ch) != "Mn")


def cell_to_str(value: object | None) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    if text.lower() in {"none", "nan"}:
        return ""
    return text


def parse_mva(raw: object | None) -> tuple[float | None, str | None, str]:
    if raw is None or raw == "":
        return None, None, "skip"
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        number = float(raw)
        if 0 < number <= 1:
            number = round(number * 100, 4)
        return number, str(raw), "numeric"

    text = str(raw).strip()
    if not text:
        return None, None, "skip"
    folded = _strip_accents(text).lower()
    if folded in {"nao", "não"}:
        return None, text, "skip"
    if any(tag in folded for tag in ("#n/d", "#n/a", "#nd")) or folded.startswith("sim"):
        return None, text, "analise"
    cleaned = text.replace("%", "").strip()
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        number = float(cleaned)
    except ValueError:
        return None, text, "analise"
    if 0 < number <= 1:
        number = round(number * 100, 4)
    return number, text, "numeric"


def classify_situacao(situacao: str, cst_saida: str, cfop: str) -> str:
    sit = _strip_accents((situacao or "").upper())
    cst_s = (cst_saida or "").strip()
    cfop_s = (cfop or "").strip()
    if "ST INTERNO" in sit:
        return "ST_INTERNO"
    if "ST NACIONAL" in sit:
        return "ST_NACIONAL"
    if "REDUC" in sit:
        return "REDUCAO"
    if "REGRA GERAL" in sit:
        return "REGRA_GERAL"
    if not cst_s or not cfop_s:
        return "INCOMPLETA"
    if cst_s in {"0", "00"} and cfop_s == "5102":
        return "REGRA_GERAL"
    if cst_s == "60" and cfop_s == "5405":
        return "ST_NACIONAL"
    if cst_s == "10" and cfop_s == "5403":
        return "ST_INTERNO"
    return "INCOMPLETA"


def _destinos_from_cells(cells: list[str], start: int) -> dict[str, str | None]:
    destinos: dict[str, str | None] = {}
    for i, key in enumerate(DESTINO_KEYS):
        raw = (cells[start + i] if start + i < len(cells) else "").strip()
        destinos[key] = raw if raw else None
    return destinos


def _rule_dict(
    *,
    source_file: str,
    source_sheet: str,
    company: str,
    ncm_original: str,
    segmento: str,
    cst_entrada: str | None,
    cst_saida: str | None,
    cfop_saida: str | None,
    destinos: dict[str, str | None],
    situacao: str,
    mva_raw: object | None,
) -> dict:
    mva_pct, mva_texto, mva_kind = parse_mva(mva_raw)
    return {
        "company": company,
        "sourceFile": source_file,
        "sourceSheet": source_sheet,
        "ncm": normalize_ncm(ncm_original),
        "ncmOriginal": str(ncm_original).strip(),
        "segmento": segmento,
        "cstEntrada": cst_entrada,
        "cstSaida": cst_saida,
        "cfopSaida": cfop_saida,
        "destinosCst": destinos,
        "situacao": situacao,
        "situacaoCodigo": classify_situacao(situacao, cst_saida or "", cfop_saida or ""),
        "mvaPercentual": mva_pct,
        "mvaTexto": mva_texto,
        "mvaKind": mva_kind,
        "observacao": None,
    }


def _payload(company: str, source: str, sheet: str, rules: list[dict], ignored: list[str]) -> dict:
    counts: dict[str, int] = {}
    for rule in rules:
        code = rule["situacaoCodigo"]
        counts[code] = counts.get(code, 0) + 1
    return {
        "company": company,
        "source": source,
        "sheet": sheet,
        "extractedSheets": [sheet],
        "ignoredSheets": ignored,
        "totalRules": len(rules),
        "uniqueNcm": len({r["ncm"] for r in rules if r["ncm"]}),
        "counts": counts,
        "rules": rules,
    }


def extract_ok_xlsx(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        first_name = wb.sheetnames[0]
        ignored = list(wb.sheetnames[1:])
        ws = wb[first_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        raise RuntimeError("OK.xlsx primeira aba vazia.")
    rules: list[dict] = []
    for raw in rows[1:]:
        cells = [cell_to_str(c) for c in (raw or ())]
        while len(cells) < 15:
            cells.append("")
        ncm_original = cells[0]
        if not ncm_original:
            continue
        if len(normalize_ncm(ncm_original)) != 8:
            continue
        mva_raw = raw[14] if raw and len(raw) > 14 else None
        rules.append(
            _rule_dict(
                source_file=path.name,
                source_sheet=first_name,
                company="baifer",
                ncm_original=ncm_original,
                segmento=cells[1],
                cst_entrada=cells[2] or None,
                cst_saida=cells[3] or None,
                cfop_saida=cells[4] or None,
                destinos=_destinos_from_cells(cells, 5),
                situacao=cells[13],
                mva_raw=mva_raw,
            )
        )
    return _payload("baifer", path.name, first_name, rules, ignored)


def _cell_text_ods(cell: TableCell) -> str:
    texts = [str(p) for p in cell.getElementsByType(P)]
    val = " ".join(texts).strip()
    if val:
        return val
    for key, value in cell.attributes.items():
        if str(key).endswith("}value") or str(key).endswith(":value"):
            if value not in (None, ""):
                return str(value)
    return ""


def _expand_row(row: TableRow, max_cols: int = 16) -> list[str]:
    out: list[str] = []
    for cell in row.getElementsByType(TableCell):
        repeated = cell.getAttribute("numbercolumnsrepeated")
        n = int(repeated) if repeated else 1
        val = _cell_text_ods(cell)
        for _ in range(n):
            out.append(val)
            if len(out) >= max_cols:
                return out
    while len(out) < max_cols:
        out.append("")
    return out[:max_cols]


def extract_ods_loja(path: Path) -> dict:
    doc = load(str(path))
    tables = doc.spreadsheet.getElementsByType(Table)
    names = [str(t.getAttribute("name")) for t in tables]
    loja = None
    for table in tables:
        name = str(table.getAttribute("name"))
        if name in FORBIDDEN_SHEETS:
            continue
        if name.strip().upper() == "LOJA":
            loja = table
            break
    if loja is None:
        raise RuntimeError("Aba LOJA não encontrada no ODS.")

    rules: list[dict] = []
    for row in loja.getElementsByType(TableRow)[1:]:
        cells = _expand_row(row, 14)
        ncm_original = (cells[0] or "").strip()
        if not ncm_original or len(normalize_ncm(ncm_original)) != 8:
            continue
        destinos = _destinos_from_cells(cells, 4)
        cst_saida = destinos.get("atacado") or destinos.get("revenda") or destinos.get("contribuinte")
        rules.append(
            _rule_dict(
                source_file=path.name,
                source_sheet="LOJA",
                company="loja",
                ncm_original=ncm_original,
                segmento=cells[1],
                cst_entrada=(cells[2] or "").strip() or None,
                cst_saida=cst_saida,
                cfop_saida=(cells[3] or "").strip() or None,
                destinos=destinos,
                situacao=(cells[12] or "").strip(),
                mva_raw=None,
            )
        )
    ignored = [n for n in names if n.strip().upper() != "LOJA"]
    return _payload("loja", path.name, "LOJA", rules, ignored)


def write_json(payload: dict, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def assert_not_mixed(baifer: dict, loja: dict) -> None:
    if baifer["company"] != "baifer" or loja["company"] != "loja":
        raise RuntimeError("company slug misturado.")
    if any(r.get("company") != "baifer" or r.get("sourceSheet") == "LOJA" for r in baifer["rules"]):
        raise RuntimeError("LOJA vazou para JSON BAIFER.")
    if any(r.get("company") != "loja" or r.get("sourceSheet") != "LOJA" for r in loja["rules"]):
        raise RuntimeError("Não-LOJA vazou para JSON LOJA.")
    if "Planilha_Classes_Fiscais" in baifer["extractedSheets"] or "Planilha_Classes_Fiscais" in loja["extractedSheets"]:
        raise RuntimeError("Aba Classes Fiscais não pode ser extraída.")


def main(argv: list[str] | None = None) -> int:
    args = argv if argv is not None else sys.argv[1:]
    ok_path = Path(args[0]) if args else DEFAULT_OK
    ods_path = Path(args[1]) if len(args) > 1 else DEFAULT_ODS
    if not ok_path.exists():
        print(f"OK.xlsx não encontrado: {ok_path}", file=sys.stderr)
        return 1
    if not ods_path.exists():
        print(f"ODS não encontrado: {ods_path}", file=sys.stderr)
        return 1
    baifer = extract_ok_xlsx(ok_path)
    loja = extract_ods_loja(ods_path)
    assert_not_mixed(baifer, loja)
    write_json(baifer, DEFAULT_BAIFER_JSON)
    write_json(loja, DEFAULT_LOJA_JSON)
    print(f"BAIFER {baifer['source']}/{baifer['sheet']}: {baifer['totalRules']} regras {baifer['counts']}")
    print(f"LOJA {loja['source']}/{loja['sheet']}: {loja['totalRules']} regras {loja['counts']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
